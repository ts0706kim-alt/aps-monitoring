# -*- coding: utf-8 -*-
"""
APS 모니터링 웹 앱 - Playwright 기반 모니터링 & 엑셀 다운로드
"""
import sys
import os
import math
from io import BytesIO
from datetime import datetime

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

from flask import Flask, render_template, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from playwright_monitor import (
    run_monitor,
    load_targets_from_csv,
    OUTPUT_COLUMNS,
)

app = Flask(__name__)
app.config["JSON_AS_ASCII"] = False

# 마지막 모니터링 결과 저장 (엑셀 다운로드용) - 10컬럼만
_last_results: list[dict] = []


def _sanitize_for_json(val):
    """NaN, inf 등을 JSON 호환 값으로 변환"""
    if val is None:
        return None
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return None
    try:
        import pandas as pd
        if pd.isna(val):
            return None
    except Exception:
        pass
    return val


def _format_price(price, currency):
    """가격 표시 포맷"""
    if price is None:
        return "-"
    c = (currency or "").strip() if isinstance(currency, str) else ""
    if c == "USD":
        return f"${price:,.2f}" if isinstance(price, (int, float)) else str(price)
    if c == "GBP":
        return f"£{price:,.2f}" if isinstance(price, (int, float)) else str(price)
    if c == "EUR":
        return f"€{price:,.2f}" if isinstance(price, (int, float)) else str(price)
    return f"{price} {c}" if c else str(price)


def _prepare_display_data(results: list[dict]) -> list[dict]:
    """웹 표시용 데이터 (10컬럼) - NaN 제거하여 JSON 호환"""
    out = []
    for r in results:
        price = _sanitize_for_json(r.get("price"))
        currency = _sanitize_for_json(r.get("currency"))
        rating = _sanitize_for_json(r.get("rating"))
        review_count = _sanitize_for_json(r.get("review_count"))
        out.append({
            "date": r.get("date") or "",
            "country": _sanitize_for_json(r.get("country")) or "-",
            "channel": _sanitize_for_json(r.get("channel")) or "-",
            "product_name": _sanitize_for_json(r.get("product_name")) or "-",
            "final_url": _sanitize_for_json(r.get("final_url")) or _sanitize_for_json(r.get("url")) or "",
            "price": _format_price(price, currency),
            "currency": (currency or "-") if isinstance(currency, str) else "-",
            "rating": rating if rating is not None else "-",
            "review_count": review_count if review_count is not None else "-",
            "promo_text": _sanitize_for_json(r.get("promo_text")) or "-",
        })
    return out


@app.route("/")
def index():
    """메인 페이지 - 모니터링 UI"""
    return render_template("index.html")


@app.route("/api/run", methods=["POST"])
def run_monitoring_api():
    """모니터링 실행 API (Playwright)"""
    global _last_results
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        os.chdir(script_dir)
        csv_path = "targets.csv" if os.path.exists("targets.csv") else "config.csv"
        if not os.path.exists(csv_path):
            return jsonify({"ok": False, "error": "targets.csv 또는 config.csv가 없습니다."}), 400

        targets = load_targets_from_csv(csv_path)
        df = run_monitor(targets, save_excel_path=None)

        cols = [c for c in OUTPUT_COLUMNS if c in df.columns]
        df_out = df[cols] if cols else df
        raw = df_out.to_dict("records")
        _last_results = [{k: _sanitize_for_json(v) for k, v in row.items()} for row in raw]

        display = _prepare_display_data(_last_results)
        return jsonify({"ok": True, "data": display})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/download/excel")
def download_excel():
    """엑셀 다운로드 (10컬럼)"""
    global _last_results
    if not _last_results:
        return jsonify({"ok": False, "error": "모니터링을 먼저 실행해 주세요."}), 400

    headers = ["date", "country", "channel", "product_name", "final_url", "price", "currency", "rating", "review_count", "promo_text"]
    header_labels = ["Date", "Country", "Channel", "Product Name", "Final URL", "Price", "Currency", "Rating", "Review Count", "Promo Text"]

    wb = Workbook()
    ws = wb.active
    ws.title = "모니터링 결과"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(header_labels, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row_idx, r in enumerate(_last_results, 2):
        for col_idx, key in enumerate(headers, 1):
            val = r.get(key)
            if val is None or (isinstance(val, float) and val != val):
                val = ""
            if key == "price" and isinstance(val, (int, float)):
                val = _format_price(val, r.get("currency"))
            ws.cell(row=row_idx, column=col_idx, value=val)
            ws.cell(row=row_idx, column=col_idx).border = thin_border

    col_widths = [12, 8, 14, 45, 55, 12, 10, 8, 12, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 50)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"aps_monitoring_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    print("=" * 50)
    print("APS 모니터링 웹 앱")
    print("브라우저에서 http://127.0.0.1:5000 접속")
    print("=" * 50)
    app.run(host="127.0.0.1", port=5000, debug=True)
